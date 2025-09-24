# excel_processing.py

from typing import List
from openpyxl import load_workbook
from models.element_infor import Story, RectColumn, Wall, CouplingBeam, Slab, Concrete, CircColumn


# ---------- STORY ----------
def read_story_table(path: str, sheet: str = "Story") -> List[Story]:
    """
    Reads story data from an Excel sheet. Assumes stories are listed
    from top to bottom (e.g., L44, L43, ... L1).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]

    stories: List[Story] = []
    # Start from the third row to skip headers
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:  # Stop if the level name is blank
            continue
        stories.append(Story(
            level=str(row[0]),
            height=float(row[1])
        ))

    # IMPORTANT: Reverse the list because ETABS requires stories
    # to be defined from the bottom up (e.g., L1, L2, L3...)
    stories.reverse()

    print(
        f"✅ Read {len(stories)} stories from '{sheet}'. First story: '{stories[0].level}', Last story: '{stories[-1].level}'.")
    return stories


# ---------- STORY ----------
def read_concrete_table(path: str, sheet_name: str = "Material") -> List[Concrete]:
    """
    Reads the Concrete material table from Excel.

    Expected format in the sheet:
        Concrete   f'c (psi)   E (lb/in2)
        5750 psi   5750        4322239.003
        ...

    :param path: Path to Excel file
    :param sheet_name: Name of sheet containing material table
    :return: List of Concrete dataclass objects
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    concretes: List[Concrete] = []

    # Skip header rows (assume first 2 rows are headers)
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        name = str(row[0]).strip()
        fc = float(row[1])
        Ec = float(row[2])
        concretes.append(Concrete(name=name, fc=fc, Ec=Ec))

    return concretes


# ----------RECTANGULAR COLUMN ----------
def read_rectangular_column_table(path: str, sheet: str = "Rectangular column") -> List[RectColumn]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]

    columns: List[RectColumn] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        col = RectColumn(
            level=row[0],
            name=row[1],
            material=row[2],
            b=float(row[4]),  # b
            h=float(row[5]),  # h
            long_bar_mat=row[6],
            tie_bar_mat=row[7],
            cover=float(row[8]),
            bars_2dir=int(row[9]),
            bars_3dir=int(row[10]),
            long_bar_size=row[11],
            tie_bar_size=row[12],
            tie_spacing=float(row[13]),
            tie_legs_2dir=int(row[14]),
            tie_legs_3dir=int(row[15])
        )
        columns.append(col)
    return columns


# ----------CIRCULAR COLUMN ----------
def read_circular_column_table(path: str, sheet: str = "Circular column") -> List[CircColumn]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]

    columns: List[CircColumn] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        col = CircColumn(
            level=row[0],
            name=row[1],
            material=row[2],
            dia=float(row[4]),  # dia
            long_bar_mat=row[5],
            tie_bar_mat=row[6],
            cover=float(row[7]),
            num_C_bars=int(row[8]),
            long_bar_size=row[9],
            tie_bar_size=row[10],
            tie_spacing=float(row[11]),
        )
        columns.append(col)
    return columns


# ---------- WALL ----------
def read_wall_table(path: str, sheet: str = "Wall") -> List[Wall]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]

    walls: List[Wall] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        level = row[0]
        material = row[1]  # "6000 psi"
        # Skip f'c column (row[2])
        name_x = row[3]
        thk_y = float(row[4])
        name_y = row[5]
        thk_x = float(row[6])

        walls.append(Wall(level=level, name=name_x, material=material, wall_thk=thk_y))
        walls.append(Wall(level=level, name=name_y, material=material, wall_thk=thk_x))

    return walls


# ---------- COUPLING BEAM ----------
def read_coupling_beam_table(path: str, sheet: str = "Coupling Beam") -> List[CouplingBeam]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]

    beams: List[CouplingBeam] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        beams.append(CouplingBeam(
            level=row[0],
            material=row[1],
            fc=int(row[2]),
            name_x=row[3],
            b_x=int(row[4]),
            h_x=int(row[5]),
            name_y=row[6],
            b_y=int(row[7]),
            h_y=int(row[8])
        ))
    return beams


# ---------- SLAB ----------
def read_slab_table(path: str, sheet: str = "Slab") -> List[Slab]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]

    slabs: List[Slab] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        slabs.append(Slab(
            level=row[0],
            material=row[1],
            fc=int(row[2]),
            name=row[3],
            thickness=float(row[4]),
            sdl=float(row[5]),
            live=float(row[6])
        ))
    return slabs
